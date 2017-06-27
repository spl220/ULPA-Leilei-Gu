from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from .models import (Language, LanguageCategory)


def export_xlsx(modeladmin, request, queryset):

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=language_export.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Languages"

    row_num = 0

    columns = [
        (u"Name", 70),
        (u"Alternative Name", 70),
        (u"Individual Language", 70),
        (u"Description", 70),
        (u"How Widely Taught", 70),
        (u"Abs Data", 70),
        (u"Abs Classifciation", 70),
        (u"Category Name", 70),
        (u"Category Description", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        category_name = ""
        category_description = ""
        language = Language.objects.prefetch_related('categories').get(pk=obj.pk)

        for category in language.categories.all():
            category_name = category.name
            category_description = category.description

        row_num += 1
        row = [
            obj.name,
            obj.alternative_name,
            obj.individual_language,
            obj.description,
            obj.how_widely_taught,
            obj.abs_data,
            obj.abs_classification,
            category_name,
            category_description,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response


class LanguageAdmin(admin.ModelAdmin):
    model = Language
    prepopulated_fields = {"slug": ("name",)}
    actions = [export_xlsx]

    list_display = ('name', 'created', 'modified')


class LanguageCategoryAdmin(admin.ModelAdmin):
    model = LanguageCategory
    prepopulated_fields = {"slug": ("name",)}

    list_display = ('name', 'created', 'modified')


admin.site.register(Language, LanguageAdmin)
admin.site.register(LanguageCategory, LanguageCategoryAdmin)

