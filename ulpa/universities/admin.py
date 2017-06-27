from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from .models import (University)


def export_xlsx(modeladmin, request, queryset):
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=university_export.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Universities"

    row_num = 0

    columns = [
        (u"Name", 70),
        (u"URL", 70),
        (u"Cross Institutional URL", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:

        row_num += 1
        row = [
            obj.name,
            obj.url,
            obj.cross_institutional_url,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True


    wb.save(response)
    return response


class UniversityAdmin(admin.ModelAdmin):
    model = University
    actions = [export_xlsx]
    list_display = ('name', 'created', 'modified')


admin.site.register(University, UniversityAdmin)
