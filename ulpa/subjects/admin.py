from django.contrib import admin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from .models import Subject, SubjectBulkUpload
from ulpa.universities.models import University
from ulpa.languages.models import Language


def export_xlsx(modeladmin, request, queryset):

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subject_export.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Universities"

    row_num = 0

    columns = [
        (u"Title", 100),
        (u"Code", 30),
        (u"URL", 70),
        (u"University", 70),
        (u"State", 70),
        (u"Other University", 70),
        (u"Intensity", 70),
        (u"Language", 70),
        (u"Prerequisite", 70),
        (u"Non-beginner Level Available", 70),
        (u"Next Offered", 70),
        (u"Study Choice", 70),
        (u"Notes", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        intensity = ';'.join(obj.intensity)
        study_choice = ';'.join(obj.study_choice)

        if ';' in intensity:
            intensity = Subject.REGULAR_AND_INTENSIVE
        if ';' in study_choice:
            study_choice = Subject.BOTH

        row = [
            obj.title,
            obj.code,
            obj.url,
            obj.university.name,
            obj.state,
            obj.other_university,
            intensity,
            obj.language.name,
            obj.prerequisite,
            obj.non_beginner_level_available,
            obj.next_offered,
            study_choice,
            obj.notes,
        ]

        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response


class SubjectAdmin(admin.ModelAdmin):
    model = Subject
    actions = [export_xlsx]

    list_display = ('title', 'university', 'language',
                    'created', 'modified',)


class SubjectBulkUploadAdmin(admin.ModelAdmin):
    model = SubjectBulkUpload

    list_display = ('id', 'status', 'user',)

admin.site.register(Subject, SubjectAdmin)
admin.site.register(SubjectBulkUpload, SubjectBulkUploadAdmin)
